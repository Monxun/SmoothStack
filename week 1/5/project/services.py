from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from itertools import islice
from datetime import datetime
import pandas as pd
import streamlit as st
import logging
import os

files = os.listdir('./data')
workbooks = [item for item in files if '.xlsx' in item]

logging.basicConfig(filename='log.log',filemode='w',format='%(asctime)s - %(levelname)s %(message)s',datefmt='%H:%M:%S', encoding='utf-8', level=logging.DEBUG)
logger = logging.getLogger()

months = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07',
          'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12', 'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07',
          'Aug': '08', 'Sep': '09', 'Oc': '10', 'Nov': '11', 'Dec': '12'}
years = ['2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021']

def check_file(file):
    while True:
        try:
            file = f'data/{file}'
            if os.path.exists(file):
                logger.info(f"{file} exists")
                break
        except FileNotFoundError:
            print("FileNotFound: not a valid file.")
        else:
            continue


def get_summary(ws, month_year_format):
    row = None
    for item in ws['A']:
        if month_year_format in str(item.value):
            row = item.row
            st.write(f'(Row: {row})')
    test = [ro for ro in ws.iter_rows(min_row=row, max_row=row, values_only=True)]
    new_test = [item for item in test[0][1:] if item != None]

    # create dictionary from row data
    row_data = {}
    row_data['30s_abandonment'] = f'Abandon after 30s: {round(new_test[1]*100,2)}%'
    row_data['fcr'] = f'FCR : {new_test[2]*100}0%'
    row_data['dsat'] = f'DSAT : {new_test[3]*100}0%'
    row_data['csat'] = f'CSAT : {new_test[4]*100}0%'

    return row_data
    

def get_current():
    # format month year for datetime comparison
    month = datetime.now().strftime('%m')
    year = datetime.now().year
    month_year_format = f'{year}-{month}'
    return month_year_format
    


def log_data(row_data):
    print(row_data)
    for item in row_data:
        logger.info(row_data[item])


def show_data(row_data):
    for item in row_data:
        st.write(row_data[item])