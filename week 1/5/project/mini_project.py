"""
1. Please import csv (or) openpyxl & Logging packages for this problem.
2. Program should take any filename as per the format mentioned as input.
3. Input month value from the file name where  eg :january(expedia_report_monthly_january_2018.xlsx)
4. Based on the month and year input values value print the values in logfile using logger

This is from the first tab :

Eg for January :

Calls Offered: 16,915
Abandon after 30s : 2.32%
FCR : 86.50%
DSAT :  14.20%
CSAT : 78.30%

Similarly go to "VOC Rolling MoM" tab

Grab all the values related to Jan-18 and print.

In Net Promoter Score : Promoters > 200 : good Promoters <200 : bad
			Passives > 100 : good Passives <100 : bad
			Decractors > 100 : good Decrators <100 : bad

Rest all values remain the same.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
import os
import streamlit as st
import pandas as pd
from services import check_file, log_data, show_data, get_current_month, get_summary

logging.basicConfig(filename='log.log',filemode='w',format='%(asctime)s - %(levelname)s %(message)s',datefmt='%H:%M:%S', encoding='utf-8', level=logging.DEBUG)

files = os.listdir('./data')
workbooks = [item for item in files if '.xlsx' in item]

months = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07',
          'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12', 'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07',
          'Aug': '08', 'Sep': '09', 'Oc': '10', 'Nov': '11', 'Dec': '12'}
years = ['2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021']


# streamlit title for page
st.title('Spreadsheet Mini-Project')

# streamlit button to get current month data if file available
col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    select_file_button = st.button('Select File')
with col2:
    current_month_button = st.button('Current Month')
# FINISH PAGE SELECTOR LOGIC

#####################################################
#SELECT MONTH

# workbook file selectbox 
workbook_selector = st.selectbox('Select a file', workbooks)  # replace with streamlit selector 

# if file selected display text
if workbook_selector:
    st.text(f'({workbook_selector} is selected.)')


# verify file exists / error handling
check_file(workbook_selector)

# initialize workbook
wb = load_workbook(f'data/{workbook_selector}')

# grab month and year from file selection
month_year = [item for item in workbook_selector.replace('.', '_').split("_") if item.capitalize() in months or item in years]

# display month / year
st.text(f'Month: {month_year[0].capitalize()}')
st.text(f'Year: {month_year[1]}')

# format month year for datetime comparison
month_year_format = f'{month_year[1]}-{months[month_year[0].capitalize()]}'

# grab worksheet
ws = wb['Summary Rolling MoM']

st.write('_' * 30)
st.subheader(f'Data for {month_year[0].capitalize()} - {month_year[1]}')

# get row data and return dictionary
row_data = get_summary(ws, month_year_format)

# log data
log_data(row_data)

# show data
show_data(row_data)

# grab worksheet

# iterate through rows and get data


# for row in ws.iter_rows(min_row=2, max_row=15):
#         row_id: datetime = row[0].value
#         if row_id.month == date.month and row_id.year == date.year:
#             log.info(f"{header[1].value.strip()}: {row[1].value:,}")
            
#             for i in range(2, 6):
#                 percentage = row[i].value * 100
#                 log.info(f"{header[i].value.strip()}: {percentage:.2f}%")
                
#             break

# the iter_row() is a generator
# so you gotta use list(worksheet_name.iter_row())


#####################################################
#CURRENT MONTH